#!/usr/bin/env python3
"""
Scaffolding Business Manager v3 - With Financial Tracking
Complete business management system with comprehensive financial management
"""

import sys
import os
import json
import sqlite3
import webbrowser
import threading
import base64
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# File upload configuration
UPLOAD_FOLDER = os.path.join(os.path.expanduser('~'), 'scaffolding_receipts')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Database setup
DB_PATH = os.path.join(os.path.expanduser('~'), 'scaffolding_business.db')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_database():
    """Initialize SQLite database with all required tables"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Invoices table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoiceNumber TEXT UNIQUE NOT NULL,
            clientName TEXT NOT NULL,
            clientAddress TEXT,
            clientPhone TEXT,
            date TEXT NOT NULL,
            status TEXT DEFAULT 'pending',
            items TEXT NOT NULL,
            subtotal REAL,
            vat REAL,
            vatApplied BOOLEAN DEFAULT 1,
            total REAL,
            notes TEXT,
            linkedJobId INTEGER,
            createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Inquiries table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS inquiries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT NOT NULL,
            email TEXT,
            location TEXT NOT NULL,
            status TEXT DEFAULT 'new',
            date TEXT NOT NULL,
            quoteAmount REAL,
            notes TEXT,
            linkedJobId INTEGER,
            createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Vehicles table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS vehicles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            registration TEXT UNIQUE NOT NULL,
            vehicleType TEXT DEFAULT 'car',
            ownerName TEXT NOT NULL,
            insuranceName TEXT NOT NULL,
            motDue TEXT,
            taxDue TEXT NOT NULL,
            tachoDue TEXT,
            insuranceDue TEXT NOT NULL,
            maintenanceDue TEXT,
            motActioned BOOLEAN DEFAULT 0,
            taxActioned BOOLEAN DEFAULT 0,
            tachoActioned BOOLEAN DEFAULT 0,
            insuranceActioned BOOLEAN DEFAULT 0,
            maintenanceActioned BOOLEAN DEFAULT 0,
            createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Jobs table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jobNumber TEXT UNIQUE NOT NULL,
            clientName TEXT NOT NULL,
            location TEXT NOT NULL,
            area TEXT,
            jobType TEXT,
            truck TEXT,
            driver TEXT,
            startDate TEXT,
            endDate TEXT,
            status TEXT DEFAULT 'pending',
            value REAL,
            linkedInvoiceId INTEGER,
            linkedInquiryId INTEGER,
            notes TEXT,
            createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updatedAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Financial Transactions table (NEW)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            transactionType TEXT NOT NULL,
            category TEXT NOT NULL,
            amount REAL NOT NULL,
            date TEXT NOT NULL,
            description TEXT NOT NULL,
            reference TEXT,
            receiptPath TEXT,
            linkedJobId INTEGER,
            notes TEXT,
            createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()
    print(f"✅ Database initialized at: {DB_PATH}")
    print(f"📁 Receipt folder: {UPLOAD_FOLDER}")

# ============================================================================
# FINANCIAL TRANSACTIONS API
# ============================================================================

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    """Get all transactions with optional filtering"""
    transaction_type = request.args.get('type')  # 'in' or 'out'
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    query = 'SELECT * FROM transactions WHERE 1=1'
    params = []
    
    if transaction_type:
        query += ' AND transactionType=?'
        params.append(transaction_type)
    
    if start_date:
        query += ' AND date >= ?'
        params.append(start_date)
    
    if end_date:
        query += ' AND date <= ?'
        params.append(end_date)
    
    query += ' ORDER BY date DESC, createdAt DESC'
    
    cursor.execute(query, params)
    transactions = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(transactions)

@app.route('/api/transactions', methods=['POST'])
def create_transaction():
    """Create a new financial transaction with optional receipt upload"""
    try:
        # Handle multipart form data
        transaction_type = request.form.get('transactionType')
        category = request.form.get('category')
        amount = float(request.form.get('amount'))
        date = request.form.get('date')
        description = request.form.get('description')
        reference = request.form.get('reference', '')
        linked_job_id = request.form.get('linkedJobId')
        notes = request.form.get('notes', '')
        
        # Handle file upload
        receipt_path = None
        if 'receipt' in request.files:
            file = request.files['receipt']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                receipt_path = filename
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO transactions (transactionType, category, amount, date, description, 
                                    reference, receiptPath, linkedJobId, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (transaction_type, category, amount, date, description, reference, 
              receipt_path, linked_job_id, notes))
        conn.commit()
        transaction_id = cursor.lastrowid
        conn.close()
        
        return jsonify({
            'id': transaction_id, 
            'message': 'Transaction created successfully',
            'receiptPath': receipt_path
        }), 201
        
    except Exception as e:
        return jsonify({'error': f'Error creating transaction: {str(e)}'}), 400

@app.route('/api/transactions/<int:transaction_id>', methods=['PUT'])
def update_transaction(transaction_id):
    """Update an existing transaction"""
    try:
        transaction_type = request.form.get('transactionType')
        category = request.form.get('category')
        amount = float(request.form.get('amount'))
        date = request.form.get('date')
        description = request.form.get('description')
        reference = request.form.get('reference', '')
        linked_job_id = request.form.get('linkedJobId')
        notes = request.form.get('notes', '')
        
        # Get existing transaction to check for receipt
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT receiptPath FROM transactions WHERE id=?', (transaction_id,))
        existing = cursor.fetchone()
        
        receipt_path = existing['receiptPath'] if existing else None
        
        # Handle new file upload
        if 'receipt' in request.files:
            file = request.files['receipt']
            if file and file.filename and allowed_file(file.filename):
                # Delete old receipt if exists
                if receipt_path:
                    old_path = os.path.join(app.config['UPLOAD_FOLDER'], receipt_path)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                
                # Save new receipt
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                receipt_path = filename
        
        cursor.execute('''
            UPDATE transactions 
            SET transactionType=?, category=?, amount=?, date=?, description=?,
                reference=?, receiptPath=?, linkedJobId=?, notes=?
            WHERE id=?
        ''', (transaction_type, category, amount, date, description, reference,
              receipt_path, linked_job_id, notes, transaction_id))
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Transaction updated successfully'})
        
    except Exception as e:
        return jsonify({'error': f'Error updating transaction: {str(e)}'}), 400

@app.route('/api/transactions/<int:transaction_id>', methods=['DELETE'])
def delete_transaction(transaction_id):
    """Delete a transaction and its receipt file"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get receipt path before deleting
    cursor.execute('SELECT receiptPath FROM transactions WHERE id=?', (transaction_id,))
    transaction = cursor.fetchone()
    
    if transaction and transaction['receiptPath']:
        receipt_path = os.path.join(app.config['UPLOAD_FOLDER'], transaction['receiptPath'])
        if os.path.exists(receipt_path):
            os.remove(receipt_path)
    
    cursor.execute('DELETE FROM transactions WHERE id=?', (transaction_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Transaction deleted successfully'})

@app.route('/api/transactions/receipts/<filename>', methods=['GET'])
def get_receipt(filename):
    """Serve uploaded receipt files"""
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except:
        return jsonify({'error': 'Receipt not found'}), 404

@app.route('/api/financial-summary', methods=['GET'])
def get_financial_summary():
    """Get financial summary with revenue, expenses, and profit"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    year = request.args.get('year')
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Build date filter
    date_filter = ''
    params = []
    
    if year:
        date_filter = " AND strftime('%Y', date) = ?"
        params.append(year)
    elif start_date and end_date:
        date_filter = " AND date BETWEEN ? AND ?"
        params.extend([start_date, end_date])
    
    # Calculate total revenue (money in)
    cursor.execute(f'''
        SELECT COALESCE(SUM(amount), 0) as total
        FROM transactions 
        WHERE transactionType = 'in' {date_filter}
    ''', params)
    revenue = cursor.fetchone()[0]
    
    # Calculate total expenses (money out)
    cursor.execute(f'''
        SELECT COALESCE(SUM(amount), 0) as total
        FROM transactions 
        WHERE transactionType = 'out' {date_filter}
    ''', params)
    expenses = cursor.fetchone()[0]
    
    # Get revenue from completed jobs
    cursor.execute(f'''
        SELECT COALESCE(SUM(value), 0) as total
        FROM jobs 
        WHERE status = 'completed' AND value IS NOT NULL {date_filter.replace('date', 'startDate')}
    ''', params)
    jobs_revenue = cursor.fetchone()[0]
    
    # Get category breakdown for expenses
    cursor.execute(f'''
        SELECT category, SUM(amount) as total, COUNT(*) as count
        FROM transactions 
        WHERE transactionType = 'out' {date_filter}
        GROUP BY category
        ORDER BY total DESC
    ''', params)
    expense_categories = [{'category': row[0], 'total': row[1], 'count': row[2]} 
                         for row in cursor.fetchall()]
    
    # Get monthly breakdown if year specified
    monthly_data = []
    if year:
        for month in range(1, 13):
            month_str = f"{year}-{month:02d}"
            cursor.execute('''
                SELECT 
                    COALESCE(SUM(CASE WHEN transactionType='in' THEN amount ELSE 0 END), 0) as revenue,
                    COALESCE(SUM(CASE WHEN transactionType='out' THEN amount ELSE 0 END), 0) as expenses
                FROM transactions 
                WHERE strftime('%Y-%m', date) = ?
            ''', (month_str,))
            row = cursor.fetchone()
            monthly_data.append({
                'month': month_str,
                'revenue': row[0],
                'expenses': row[1],
                'profit': row[0] - row[1]
            })
    
    conn.close()
    
    total_revenue = revenue + jobs_revenue
    profit = total_revenue - expenses
    
    return jsonify({
        'revenue': total_revenue,
        'manual_revenue': revenue,
        'jobs_revenue': jobs_revenue,
        'expenses': expenses,
        'profit': profit,
        'expense_categories': expense_categories,
        'monthly_data': monthly_data
    })

@app.route('/api/financial-report', methods=['GET'])
def generate_financial_report():
    """Generate comprehensive financial report for printing"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({'error': 'Start date and end date required'}), 400
    
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get all transactions in date range
    cursor.execute('''
        SELECT * FROM transactions 
        WHERE date BETWEEN ? AND ?
        ORDER BY date DESC, transactionType
    ''', (start_date, end_date))
    transactions = [dict(row) for row in cursor.fetchall()]
    
    # Get completed jobs in date range
    cursor.execute('''
        SELECT * FROM jobs 
        WHERE status = 'completed' AND startDate BETWEEN ? AND ?
        ORDER BY startDate DESC
    ''', (start_date, end_date))
    completed_jobs = [dict(row) for row in cursor.fetchall()]
    
    # Calculate totals
    revenue_in = sum(t['amount'] for t in transactions if t['transactionType'] == 'in')
    expenses_out = sum(t['amount'] for t in transactions if t['transactionType'] == 'out')
    jobs_revenue = sum(j['value'] or 0 for j in completed_jobs)
    total_revenue = revenue_in + jobs_revenue
    profit = total_revenue - expenses_out
    
    conn.close()
    
    return jsonify({
        'period': {
            'start': start_date,
            'end': end_date
        },
        'transactions': transactions,
        'completed_jobs': completed_jobs,
        'totals': {
            'revenue': total_revenue,
            'manual_revenue': revenue_in,
            'jobs_revenue': jobs_revenue,
            'expenses': expenses_out,
            'profit': profit
        }
    })

# ============================================================================
# EXISTING API ROUTES (Invoices, Inquiries, Jobs, Vehicles)
# ============================================================================

# API Routes - Invoices
@app.route('/api/invoices', methods=['GET'])
def get_invoices():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM invoices ORDER BY date DESC')
    invoices = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(invoices)

@app.route('/api/invoices', methods=['POST'])
def create_invoice():
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        vat_applied = data.get('vatApplied', True)
        vat = data.get('vat', 0)
        if not vat_applied:
            vat = 0
        
        cursor.execute('''
            INSERT INTO invoices (invoiceNumber, clientName, clientAddress, clientPhone, 
                                date, status, items, subtotal, vat, vatApplied, total, notes, linkedJobId)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['invoiceNumber'], data['clientName'], data.get('clientAddress'),
            data.get('clientPhone'), data['date'], data['status'], data['items'],
            data['subtotal'], vat, vat_applied, data['total'], data.get('notes'),
            data.get('linkedJobId')
        ))
        conn.commit()
        invoice_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': invoice_id, 'message': 'Invoice created successfully'}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Invoice number already exists'}), 400
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Error creating invoice: {str(e)}'}), 400

@app.route('/api/invoices/<int:invoice_id>', methods=['PUT'])
def update_invoice(invoice_id):
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    vat_applied = data.get('vatApplied', True)
    vat = data.get('vat', 0)
    if not vat_applied:
        vat = 0
    
    cursor.execute('''
        UPDATE invoices 
        SET invoiceNumber=?, clientName=?, clientAddress=?, clientPhone=?,
            date=?, status=?, items=?, subtotal=?, vat=?, vatApplied=?, total=?, notes=?, linkedJobId=?
        WHERE id=?
    ''', (
        data['invoiceNumber'], data['clientName'], data.get('clientAddress'),
        data.get('clientPhone'), data['date'], data['status'], data['items'],
        data['subtotal'], vat, vat_applied, data['total'], data.get('notes'), 
        data.get('linkedJobId'), invoice_id
    ))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Invoice updated successfully'})

@app.route('/api/invoices/<int:invoice_id>', methods=['DELETE'])
def delete_invoice(invoice_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM invoices WHERE id=?', (invoice_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Invoice deleted successfully'})

@app.route('/api/invoices/<int:invoice_id>/preview', methods=['GET'])
def preview_invoice(invoice_id):
    """Generate professional HTML invoice preview - single page, print-ready"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM invoices WHERE id=?', (invoice_id,))
        invoice_row = cursor.fetchone()
        conn.close()
        
        if not invoice_row:
            return jsonify({'error': 'Invoice not found'}), 404
        
        invoice = dict(invoice_row)
        items = json.loads(invoice['items'])
        
        invoice_date = datetime.strptime(invoice['date'], '%Y-%m-%d').strftime('%d/%m/%Y')
        
        # Build items HTML
        items_html = ''
        for idx, item in enumerate(items, 1):
            amount = item.get('quantity', 1) * item.get('rate', 0)
            items_html += f'''
                <tr>
                    <td style="width: 40px; text-align: center; padding: 8px;">{idx}</td>
                    <td style="padding: 8px;">{item['description']}</td>
                    <td style="text-align: center; padding: 8px;">{item.get('quantity', 1)}</td>
                    <td style="text-align: right; padding: 8px;">£{item.get('rate', 0):,.0f}</td>
                    <td style="text-align: right; padding: 8px;">£{amount:,.0f}</td>
                </tr>
'''
        
        html = f'''
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Invoice {invoice['invoiceNumber']} - Khalsa Scaffolding LTD</title>
    <style>
        @page {{
            margin: 15mm;
            size: A4 portrait;
        }}
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Calibri', 'Arial', sans-serif;
            font-size: 10pt;
            color: #000;
            background: #f5f5f5;
            padding: 15px;
        }}
        .page-wrapper {{
            max-width: 210mm;
            margin: 0 auto;
            background: white;
            box-shadow: 0 0 15px rgba(0,0,0,0.1);
        }}
        .invoice-container {{
            padding: 20px 30px;
            background: white;
        }}
        
        /* Action Buttons */
        .action-bar {{
            background: #4472C4;
            padding: 12px 30px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            color: white;
        }}
        .action-bar h2 {{
            font-size: 14pt;
            font-weight: normal;
        }}
        .action-btn {{
            background: white;
            color: #4472C4;
            border: none;
            padding: 8px 18px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 10pt;
            font-weight: bold;
            transition: all 0.2s;
        }}
        .action-btn:hover {{
            background: #f0f0f0;
            transform: translateY(-1px);
        }}
        
        /* Header */
        .header {{
            display: flex;
            justify-content: space-between;
            margin-bottom: 20px;
            padding-bottom: 12px;
            border-bottom: 2px solid #4472C4;
        }}
        .company-info {{
            flex: 1;
        }}
        .company-name {{
            font-size: 16pt;
            font-weight: bold;
            color: #000;
            margin-bottom: 4px;
        }}
        .company-details {{
            font-size: 9pt;
            line-height: 1.4;
            color: #333;
        }}
        .invoice-title {{
            text-align: right;
        }}
        .invoice-title h1 {{
            font-size: 28pt;
            font-weight: bold;
            color: #000;
            margin-bottom: 0;
        }}
        
        /* Invoice Meta */
        .invoice-meta {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
            margin: 15px 0;
        }}
        .meta-section {{
            background: #fafafa;
            padding: 12px;
            border-left: 3px solid #4472C4;
        }}
        .meta-section h3 {{
            font-size: 9pt;
            font-weight: bold;
            margin-bottom: 6px;
            color: #000;
            text-transform: uppercase;
        }}
        .meta-section p {{
            font-size: 9pt;
            line-height: 1.4;
            margin: 2px 0;
        }}
        .meta-label {{
            display: inline-block;
            width: 70px;
            font-weight: bold;
        }}
        
        /* Items Table */
        .items-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
        }}
        .items-table thead {{
            background: #4472C4;
            color: white;
        }}
        .items-table th {{
            padding: 8px;
            text-align: left;
            font-weight: bold;
            font-size: 9pt;
            text-transform: uppercase;
        }}
        .items-table tbody td {{
            padding: 8px;
            border-bottom: 1px solid #e0e0e0;
            font-size: 9pt;
        }}
        .items-table tbody tr:last-child td {{
            border-bottom: 2px solid #4472C4;
        }}
        
        /* Totals */
        .totals-section {{
            margin-top: 15px;
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
        }}
        .thank-you {{
            flex: 1;
            font-size: 10pt;
            font-weight: bold;
        }}
        .totals-table {{
            width: 280px;
        }}
        .totals-table tr td {{
            padding: 6px 12px;
            font-size: 10pt;
        }}
        .totals-table tr td:first-child {{
            font-weight: bold;
            text-align: right;
            text-transform: uppercase;
        }}
        .totals-table tr td:last-child {{
            text-align: right;
        }}
        .totals-table .subtotal-row {{
            border-top: 1px solid #e0e0e0;
        }}
        .totals-table .tax-row {{
            color: #666;
            font-size: 9pt;
        }}
        .totals-table .total-row {{
            background: #4472C4;
            color: white;
            font-weight: bold;
            font-size: 12pt;
        }}
        .totals-table .total-row td {{
            padding: 10px 12px;
        }}
        
        /* Bank Details */
        .bank-details {{
            background: #fafafa;
            padding: 12px;
            margin-top: 15px;
            border-radius: 4px;
            border: 1px solid #e0e0e0;
        }}
        .bank-details h4 {{
            font-size: 10pt;
            font-weight: bold;
            margin-bottom: 6px;
            color: #4472C4;
        }}
        .bank-details p {{
            font-size: 9pt;
            line-height: 1.4;
            margin: 2px 0;
        }}
        .bank-label {{
            display: inline-block;
            width: 110px;
            font-weight: bold;
        }}
        
        /* Footer */
        .footer {{
            margin-top: 15px;
            padding-top: 12px;
            border-top: 1px solid #e0e0e0;
            text-align: center;
            font-size: 8pt;
            color: #666;
        }}
        .footer p {{
            margin: 3px 0;
        }}
        
        /* Print Styles */
        @media print {{
            body {{
                background: white;
                padding: 0;
            }}
            .page-wrapper {{
                box-shadow: none;
                max-width: 100%;
            }}
            .action-bar {{
                display: none;
            }}
            .invoice-container {{
                padding: 0;
            }}
        }}
    </style>
</head>
<body>
    <div class="page-wrapper">
        <div class="action-bar">
            <h2>📄 Invoice Preview</h2>
            <button class="action-btn" onclick="window.print()">
                🖨️ Print / Download PDF
            </button>
        </div>
        
        <div class="invoice-container">
            <!-- Header -->
            <div class="header">
                <div class="company-info">
                    <div class="company-name">Khalsa Sacffolding LTD</div>
                    <div class="company-details">
                        66 Raynton Drive<br>
                        Hayes UB4 8BE<br>
                        Phone: 07741252013
                    </div>
                </div>
                <div class="invoice-title">
                    <h1>INVOICE</h1>
                </div>
            </div>
            
            <!-- Invoice Meta Info -->
            <div class="invoice-meta">
                <div class="meta-section">
                    <h3>Bill To</h3>
                    <p><strong>{invoice['clientName']}</strong></p>
                    <p>{invoice.get('clientPhone', '')}</p>
                    <p>{invoice.get('clientAddress', '')}</p>
                </div>
                <div class="meta-section">
                    <h3>Invoice Details</h3>
                    <p><span class="meta-label">Invoice #:</span> {invoice['invoiceNumber']}</p>
                    <p><span class="meta-label">Date:</span> {invoice_date}</p>
                </div>
            </div>
            
            <!-- Items Table -->
            <table class="items-table">
                <thead>
                    <tr>
                        <th style="width: 40px;">#</th>
                        <th>Description</th>
                        <th style="width: 60px; text-align: center;">Qty</th>
                        <th style="width: 100px; text-align: right;">Unit Price</th>
                        <th style="width: 100px; text-align: right;">Amount</th>
                    </tr>
                </thead>
                <tbody>
                    {items_html}
                </tbody>
            </table>
            
            <!-- Totals -->
            <div class="totals-section">
                <div class="thank-you">
                    Thank you for your business!
                </div>
                <table class="totals-table">
                    <tr class="subtotal-row">
                        <td>Subtotal:</td>
                        <td>£{invoice['subtotal']:,.0f}</td>
                    </tr>'''
        
        if invoice.get('vatApplied', True):
            html += f'''
                    <tr class="tax-row">
                        <td>Tax Rate:</td>
                        <td>20%</td>
                    </tr>
                    <tr class="tax-row">
                        <td>VAT:</td>
                        <td>£{invoice['vat']:,.0f}</td>
                    </tr>'''
        
        html += f'''
                    <tr class="total-row">
                        <td>Total:</td>
                        <td>£{invoice['total']:,.0f}</td>
                    </tr>
                </table>
            </div>
            
            <!-- Bank Details -->
            <div class="bank-details">
                <h4>Payment Information</h4>
                <p><span class="bank-label">Bank name:</span> Khalsa Scaffolding LTD</p>
                <p><span class="bank-label">Account number:</span> 33189759</p>
                <p><span class="bank-label">Sort code:</span> 20-42-76</p>
            </div>
            
            <!-- Footer -->
            <div class="footer">
                <p><strong>If you have any questions about this invoice, please contact:</strong></p>
                <p>Jagtar Singh Brar, 07741252013</p>
                <p>Company number: 13490441 | VAT: 43747190</p>
            </div>
        </div>
    </div>
</body>
</html>
'''
        
        from flask import make_response
        response = make_response(html)
        response.headers['Content-Type'] = 'text/html'
        return response
    except Exception as e:
        return jsonify({'error': f'Error generating preview: {str(e)}'}), 500

@app.route('/api/invoices/<int:invoice_id>/excel', methods=['GET'])
def download_invoice_excel(invoice_id):
    """Generate Excel invoice matching the template design"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM invoices WHERE id=?', (invoice_id,))
        invoice_row = cursor.fetchone()
        conn.close()
        
        if not invoice_row:
            return jsonify({'error': 'Invoice not found'}), 404
        
        invoice = dict(invoice_row)
        items = json.loads(invoice['items'])
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Invoice {invoice['invoiceNumber']}"
        
        # Define styles
        header_font = Font(name='Calibri', size=14, bold=True, color='000000')
        invoice_title_font = Font(name='Calibri', size=18, bold=True, color='000000')
        section_header_font = Font(name='Calibri', size=11, bold=True, color='000000')
        normal_font = Font(name='Calibri', size=11, color='000000')
        small_font = Font(name='Calibri', size=9, color='000000')
        
        table_header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        table_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        # Row 1: Company Name and Invoice Title
        ws['A1'] = 'Khalsa Sacffolding LTD'
        ws['A1'].font = header_font
        ws['G1'] = 'INVOICE'
        ws['G1'].font = invoice_title_font
        ws['G1'].alignment = Alignment(horizontal='right', vertical='top')
        
        # Row 2-3: Company Address
        ws['A2'] = '66 Raynton Drive'
        ws['A2'].font = normal_font
        ws['A3'] = 'Hayes UB4 8BE'
        ws['A3'].font = normal_font
        
        # Row 4-5: Phone and Invoice Details
        ws['A4'] = 'Phone: 07741252013'
        ws['A4'].font = normal_font
        ws['F4'] = 'INVOICE #'
        ws['F4'].font = section_header_font
        ws['H4'] = 'DATE'
        ws['H4'].font = section_header_font
        
        ws['F5'] = invoice['invoiceNumber']
        ws['F5'].font = normal_font
        ws['H5'] = datetime.strptime(invoice['date'], '%Y-%m-%d').strftime('%Y-%m-%d')
        ws['H5'].font = normal_font
        
        # Row 7-10: Bill To Section
        ws['A7'] = 'BILL TO'
        ws['A7'].font = section_header_font
        ws['F7'] = 'CUSTOMER ID'
        ws['F7'].font = section_header_font
        ws['H7'] = 'TERMS'
        ws['H7'].font = section_header_font
        
        ws['A8'] = invoice['clientName']
        ws['A8'].font = normal_font
        ws['H8'] = 'Due Upon Receipt'
        ws['H8'].font = normal_font
        
        # Client contact person (if available in clientPhone field, we'll put it here)
        if invoice.get('clientPhone'):
            ws['A9'] = invoice['clientPhone']
            ws['A9'].font = normal_font
        
        # Client address
        if invoice.get('clientAddress'):
            ws['A10'] = invoice['clientAddress']
            ws['A10'].font = normal_font
        
        # Row 15: Table Headers
        current_row = 15
        ws[f'A{current_row}'] = 'DESCRIPTION'
        ws[f'F{current_row}'] = 'QTY'
        ws[f'G{current_row}'] = 'UNIT PRICE'
        ws[f'H{current_row}'] = 'AMOUNT'
        
        for col in ['A', 'F', 'G', 'H']:
            cell = ws[f'{col}{current_row}']
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.alignment = Alignment(horizontal='center' if col != 'A' else 'left', vertical='center')
            cell.border = border_thin
        
        # Add invoice items
        current_row = 16
        item_start_row = current_row
        for idx, item in enumerate(items, 1):
            ws[f'A{current_row}'] = idx
            ws[f'C{current_row}'] = item['description']
            ws[f'F{current_row}'] = item.get('quantity', 1)
            ws[f'G{current_row}'] = item.get('rate', 0)
            
            # Calculate amount
            amount = item.get('quantity', 1) * item.get('rate', 0)
            ws[f'H{current_row}'] = amount
            
            # Format cells
            ws[f'A{current_row}'].font = normal_font
            ws[f'C{current_row}'].font = normal_font
            ws[f'F{current_row}'].font = normal_font
            ws[f'F{current_row}'].alignment = Alignment(horizontal='center')
            ws[f'G{current_row}'].font = normal_font
            ws[f'G{current_row}'].number_format = '#,##0'
            ws[f'H{current_row}'].font = normal_font
            ws[f'H{current_row}'].number_format = '#,##0'
            
            current_row += 1
        
        item_end_row = current_row - 1
        
        # Add thank you message and totals section
        current_row = 23
        ws[f'A{current_row}'] = 'Thank you for your business!'
        ws[f'A{current_row}'].font = normal_font
        
        ws[f'F{current_row}'] = 'SUBTOTAL'
        ws[f'F{current_row}'].font = section_header_font
        ws[f'H{current_row}'] = f'=SUM(H{item_start_row}:H{item_end_row})'
        ws[f'H{current_row}'].number_format = '#,##0'
        
        current_row = 24
        ws[f'F{current_row}'] = 'TAX RATE'
        ws[f'F{current_row}'].font = section_header_font
        tax_rate = 0.2 if invoice.get('vatApplied', True) else 0
        ws[f'H{current_row}'] = tax_rate
        ws[f'H{current_row}'].number_format = '0.0'
        
        current_row = 25
        ws[f'B{current_row}'] = 'Bank name: Khalsa Scaffolding LTD'
        ws[f'B{current_row}'].font = normal_font
        ws[f'F{current_row}'] = 'VAT'
        ws[f'F{current_row}'].font = section_header_font
        ws[f'H{current_row}'] = '=H23*H24'
        ws[f'H{current_row}'].number_format = '#,##0'
        
        current_row = 26
        ws[f'B{current_row}'] = 'Account number: 33189759'
        ws[f'B{current_row}'].font = normal_font
        ws[f'F{current_row}'] = 'TOTAL'
        ws[f'F{current_row}'].font = Font(name='Calibri', size=11, bold=True, color='000000')
        ws[f'H{current_row}'] = '=H23+H25'
        ws[f'H{current_row}'].font = Font(name='Calibri', size=11, bold=True, color='000000')
        ws[f'H{current_row}'].number_format = '#,##0'
        
        current_row = 27
        ws[f'B{current_row}'] = 'Sort code: 20-42-76'
        ws[f'B{current_row}'].font = normal_font
        
        current_row = 29
        ws[f'C{current_row}'] = '                         company number: 13490441 VAT:43747190'
        ws[f'C{current_row}'].font = small_font
        
        current_row = 30
        ws[f'A{current_row}'] = 'If you have any questions about this invoice, please contact'
        ws[f'A{current_row}'].font = small_font
        
        current_row = 31
        ws[f'A{current_row}'] = 'Jagtar Singh Brar, 07741252013'
        ws[f'A{current_row}'].font = small_font
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Invoice_{invoice["invoiceNumber"]}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error generating Excel invoice: {str(e)}'}), 500

        
    except Exception as e:
        return jsonify({'error': f'Error generating invoice preview: {str(e)}'}), 500

# API Routes - Inquiries
@app.route('/api/inquiries', methods=['GET'])
def get_inquiries():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM inquiries ORDER BY date DESC')
    inquiries = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(inquiries)

@app.route('/api/inquiries', methods=['POST'])
def create_inquiry():
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO inquiries (name, phone, email, location, status, date, quoteAmount, notes, linkedJobId)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['name'], data['phone'], data.get('email'), data['location'],
        data.get('status', 'new'), data['date'], data.get('quoteAmount'),
        data.get('notes'), data.get('linkedJobId')
    ))
    conn.commit()
    inquiry_id = cursor.lastrowid
    conn.close()
    return jsonify({'id': inquiry_id, 'message': 'Inquiry created successfully'}), 201

@app.route('/api/inquiries/<int:inquiry_id>', methods=['PUT'])
def update_inquiry(inquiry_id):
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE inquiries 
        SET name=?, phone=?, email=?, location=?, status=?, date=?, quoteAmount=?, notes=?, linkedJobId=?
        WHERE id=?
    ''', (
        data['name'], data['phone'], data.get('email'), data['location'],
        data['status'], data['date'], data.get('quoteAmount'), data.get('notes'),
        data.get('linkedJobId'), inquiry_id
    ))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Inquiry updated successfully'})

@app.route('/api/inquiries/<int:inquiry_id>', methods=['DELETE'])
def delete_inquiry(inquiry_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM inquiries WHERE id=?', (inquiry_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Inquiry deleted successfully'})

# API Routes - Jobs
@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM jobs ORDER BY createdAt DESC')
    jobs = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(jobs)

@app.route('/api/jobs', methods=['POST'])
def create_job():
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO jobs (jobNumber, clientName, location, area, jobType, truck, driver, startDate, 
                            endDate, status, value, linkedInvoiceId, linkedInquiryId, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['jobNumber'], data['clientName'], data['location'], data.get('area'),
            data.get('jobType'), data.get('truck'), data.get('driver'), data.get('startDate'), data.get('endDate'),
            data.get('status', 'pending'), data.get('value'), data.get('linkedInvoiceId'),
            data.get('linkedInquiryId'), data.get('notes')
        ))
        conn.commit()
        job_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': job_id, 'message': 'Job created successfully'}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Job number already exists'}), 400
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Error creating job: {str(e)}'}), 400

@app.route('/api/jobs/<int:job_id>', methods=['PUT'])
def update_job(job_id):
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute('''
            UPDATE jobs 
            SET jobNumber=?, clientName=?, location=?, area=?, jobType=?, truck=?, driver=?, startDate=?,
                endDate=?, status=?, value=?, linkedInvoiceId=?, linkedInquiryId=?, notes=?,
                updatedAt=CURRENT_TIMESTAMP
            WHERE id=?
        ''', (
            data['jobNumber'], data['clientName'], data['location'], data.get('area'),
            data.get('jobType'), data.get('truck'), data.get('driver'), data.get('startDate'), data.get('endDate'),
            data.get('status'), data.get('value'), data.get('linkedInvoiceId'),
            data.get('linkedInquiryId'), data.get('notes'), job_id
        ))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Job updated successfully'})
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Error updating job: {str(e)}'}), 400

@app.route('/api/jobs/<int:job_id>', methods=['DELETE'])
def delete_job(job_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM jobs WHERE id=?', (job_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Job deleted successfully'})

@app.route('/api/jobs/bulk-delete', methods=['POST'])
def bulk_delete_jobs():
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs provided'}), 400
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(ids))
    cursor.execute(f'DELETE FROM jobs WHERE id IN ({placeholders})', ids)
    conn.commit()
    conn.close()
    return jsonify({'message': f'{len(ids)} jobs deleted successfully'})

@app.route('/api/jobs/export', methods=['GET'])
def export_jobs():
    import csv
    from io import StringIO
    
    area = request.args.get('area', 'all')
    
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    if area == 'all':
        cursor.execute('SELECT * FROM jobs ORDER BY createdAt DESC')
    else:
        cursor.execute('SELECT * FROM jobs WHERE area = ? ORDER BY createdAt DESC', (area,))
    
    jobs = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    if not jobs:
        return jsonify({'error': 'No jobs to export'}), 404
    
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=jobs[0].keys())
    writer.writeheader()
    writer.writerows(jobs)
    
    from flask import make_response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=jobs_export_{area}_{datetime.now().strftime("%Y%m%d")}.csv'
    return response

# API Routes - Vehicles
@app.route('/api/vehicles', methods=['GET'])
def get_vehicles():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM vehicles ORDER BY registration')
    vehicles = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(vehicles)

@app.route('/api/vehicles', methods=['POST'])
def create_vehicle():
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO vehicles (registration, vehicleType, ownerName, insuranceName,
                                motDue, taxDue, tachoDue, insuranceDue, maintenanceDue, 
                                motActioned, taxActioned, tachoActioned, insuranceActioned, maintenanceActioned)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['registration'].upper(), data.get('vehicleType', 'car'), data['ownerName'], data['insuranceName'],
            data.get('motDue'), data['taxDue'], data.get('tachoDue'), data['insuranceDue'], data.get('maintenanceDue'),
            data.get('motActioned', False), data.get('taxActioned', False), 
            data.get('tachoActioned', False), data.get('insuranceActioned', False), data.get('maintenanceActioned', False)
        ))
        conn.commit()
        vehicle_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': vehicle_id, 'message': 'Vehicle created successfully'}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Registration number already exists'}), 400

@app.route('/api/vehicles/<int:vehicle_id>', methods=['PUT'])
def update_vehicle(vehicle_id):
    data = request.json
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE vehicles 
        SET registration=?, vehicleType=?, ownerName=?, insuranceName=?,
            motDue=?, taxDue=?, tachoDue=?, insuranceDue=?, maintenanceDue=?, 
            motActioned=?, taxActioned=?, tachoActioned=?, insuranceActioned=?, maintenanceActioned=?
        WHERE id=?
    ''', (
        data['registration'].upper(), data.get('vehicleType', 'car'), data['ownerName'], data['insuranceName'],
        data.get('motDue'), data['taxDue'], data.get('tachoDue'), data['insuranceDue'], data.get('maintenanceDue'),
        data.get('motActioned', False), data.get('taxActioned', False),
        data.get('tachoActioned', False), data.get('insuranceActioned', False), 
        data.get('maintenanceActioned', False), vehicle_id
    ))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Vehicle updated successfully'})

@app.route('/api/vehicles/<int:vehicle_id>', methods=['DELETE'])
def delete_vehicle(vehicle_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM vehicles WHERE id=?', (vehicle_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Vehicle deleted successfully'})

# Serve the HTML interface
@app.route('/')
def index():
    return send_from_directory('.', 'complete_scaffolding_dashboard.html')

def open_browser():
    """Open the default browser after a short delay"""
    import time
    time.sleep(1.5)
    webbrowser.open('http://127.0.0.1:5000')

def main():
    """Main application entry point"""
    print("=" * 60)
    print("🏗️ KHALSA SCAFFOLDING - BUSINESS MANAGER V3")
    print("=" * 60)
    print()
    
    init_database()
    
    print()
    print("🚀 Starting server...")
    print("📍 Server will run at: http://127.0.0.1:5000")
    print("🗄️ Database location:", DB_PATH)
    print("📁 Receipts folder:", UPLOAD_FOLDER)
    print("✨ NEW: Complete Financial Tracking System!")
    print("💰 Track money in/out, upload receipts, generate reports")
    print()
    print("💡 TIP: Keep this window open while using the application")
    print("⚠️ Press Ctrl+C to stop the server")
    print()
    
    threading.Thread(target=open_browser, daemon=True).start()
    
    try:
        app.run(host='127.0.0.1', port=5000, debug=False)
    except KeyboardInterrupt:
        print("\n\n👋 Server stopped. Goodbye!")
        sys.exit(0)

if __name__ == '__main__':
    main()
