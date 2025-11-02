from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
import io
import os
import sys
import webbrowser
import threading

# Determine if running as exe or script
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
CORS(app)

# Database configuration
DB_PATH = os.path.join(BASE_DIR, 'scaffolding_business.db')
TEMPLATE_PATH = os.path.join(BASE_DIR, 'invoice_template.xlsx')

app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Database Models
class Invoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_number = db.Column(db.String(50), unique=True, nullable=False)
    client_name = db.Column(db.String(200), nullable=False)
    client_address = db.Column(db.String(500))
    client_phone = db.Column(db.String(50))
    date = db.Column(db.String(20), nullable=False)
    status = db.Column(db.String(20), default='pending')
    items = db.Column(db.Text)
    subtotal = db.Column(db.Float, default=0)
    vat = db.Column(db.Float, default=0)
    total = db.Column(db.Float, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    notes = db.Column(db.Text)

class InvoiceChange(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoice.id'), nullable=False)
    change_type = db.Column(db.String(50))
    change_description = db.Column(db.Text)
    old_value = db.Column(db.Text)
    new_value = db.Column(db.Text)
    changed_at = db.Column(db.DateTime, default=datetime.utcnow)
    changed_by = db.Column(db.String(100), default='user')

class Inquiry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    phone = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(200))
    location = db.Column(db.String(200), nullable=False)
    status = db.Column(db.String(50), default='new')
    date = db.Column(db.String(20), nullable=False)
    notes = db.Column(db.Text)
    quote_amount = db.Column(db.Float)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class InquiryChange(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    inquiry_id = db.Column(db.Integer, db.ForeignKey('inquiry.id'), nullable=False)
    change_type = db.Column(db.String(50))
    change_description = db.Column(db.Text)
    old_value = db.Column(db.Text)
    new_value = db.Column(db.Text)
    changed_at = db.Column(db.DateTime, default=datetime.utcnow)
    changed_by = db.Column(db.String(100), default='user')

class Vehicle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    registration = db.Column(db.String(20), unique=True, nullable=False)
    make = db.Column(db.String(100), nullable=False)
    model = db.Column(db.String(100), nullable=False)
    colour = db.Column(db.String(50))
    fuel_type = db.Column(db.String(50))
    mot_due = db.Column(db.String(20), nullable=False)
    tax_due = db.Column(db.String(20), nullable=False)
    insurance_due = db.Column(db.String(20), nullable=False)
    mot_actioned = db.Column(db.Boolean, default=False)
    tax_actioned = db.Column(db.Boolean, default=False)
    insurance_actioned = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class VehicleChange(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicle.id'), nullable=False)
    change_type = db.Column(db.String(50))
    change_description = db.Column(db.Text)
    changed_at = db.Column(db.DateTime, default=datetime.utcnow)

def log_change(model_class, item_id, change_type, description, old_value=None, new_value=None):
    if model_class == Invoice:
        change = InvoiceChange(
            invoice_id=item_id,
            change_type=change_type,
            change_description=description,
            old_value=old_value,
            new_value=new_value
        )
    elif model_class == Inquiry:
        change = InquiryChange(
            inquiry_id=item_id,
            change_type=change_type,
            change_description=description,
            old_value=old_value,
            new_value=new_value
        )
    elif model_class == Vehicle:
        change = VehicleChange(
            vehicle_id=item_id,
            change_type=change_type,
            change_description=description
        )
    db.session.add(change)
    db.session.commit()

# ========== INVOICE ENDPOINTS ==========

@app.route('/api/invoices', methods=['GET'])
def get_invoices():
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    year = request.args.get('year', '')
    
    query = Invoice.query
    
    if search:
        query = query.filter(
            db.or_(
                Invoice.invoice_number.ilike(f'%{search}%'),
                Invoice.client_name.ilike(f'%{search}%')
            )
        )
    
    if status:
        query = query.filter(Invoice.status == status)
    
    if year:
        query = query.filter(Invoice.date.like(f'{year}%'))
    
    invoices = query.order_by(Invoice.created_at.desc()).all()
    
    return jsonify([{
        'id': inv.id,
        'invoiceNumber': inv.invoice_number,
        'clientName': inv.client_name,
        'clientAddress': inv.client_address,
        'clientPhone': inv.client_phone,
        'date': inv.date,
        'status': inv.status,
        'items': inv.items,
        'subtotal': inv.subtotal,
        'vat': inv.vat,
        'total': inv.total,
        'notes': inv.notes,
        'createdAt': inv.created_at.isoformat(),
        'updatedAt': inv.updated_at.isoformat()
    } for inv in invoices])

@app.route('/api/invoices', methods=['POST'])
def create_invoice():
    data = request.json
    
    invoice = Invoice(
        invoice_number=data.get('invoiceNumber'),
        client_name=data.get('clientName'),
        client_address=data.get('clientAddress'),
        client_phone=data.get('clientPhone'),
        date=data.get('date'),
        status=data.get('status', 'pending'),
        items=data.get('items'),
        subtotal=data.get('subtotal', 0),
        vat=data.get('vat', 0),
        total=data.get('total', 0),
        notes=data.get('notes', '')
    )
    
    db.session.add(invoice)
    db.session.commit()
    
    log_change(Invoice, invoice.id, 'created', f'Invoice {invoice.invoice_number} created')
    
    return jsonify({'id': invoice.id, 'message': 'Invoice created successfully'}), 201

@app.route('/api/invoices/<int:invoice_id>', methods=['PUT'])
def update_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    data = request.json
    
    old_status = invoice.status
    
    invoice.client_name = data.get('clientName', invoice.client_name)
    invoice.client_address = data.get('clientAddress', invoice.client_address)
    invoice.client_phone = data.get('clientPhone', invoice.client_phone)
    invoice.date = data.get('date', invoice.date)
    invoice.status = data.get('status', invoice.status)
    invoice.items = data.get('items', invoice.items)
    invoice.subtotal = data.get('subtotal', invoice.subtotal)
    invoice.vat = data.get('vat', invoice.vat)
    invoice.total = data.get('total', invoice.total)
    invoice.notes = data.get('notes', invoice.notes)
    
    db.session.commit()
    
    if old_status != invoice.status:
        log_change(Invoice, invoice.id, 'status_changed', 
                  f'Status changed from {old_status} to {invoice.status}',
                  old_status, invoice.status)
    else:
        log_change(Invoice, invoice.id, 'updated', f'Invoice {invoice.invoice_number} updated')
    
    return jsonify({'message': 'Invoice updated successfully'})

@app.route('/api/invoices/<int:invoice_id>', methods=['DELETE'])
def delete_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    invoice_number = invoice.invoice_number
    
    db.session.delete(invoice)
    db.session.commit()
    
    return jsonify({'message': f'Invoice {invoice_number} deleted successfully'})

@app.route('/api/invoices/<int:invoice_id>/changes', methods=['GET'])
def get_invoice_changes(invoice_id):
    changes = InvoiceChange.query.filter_by(invoice_id=invoice_id).order_by(InvoiceChange.changed_at.desc()).all()
    
    return jsonify([{
        'id': change.id,
        'changeType': change.change_type,
        'description': change.change_description,
        'oldValue': change.old_value,
        'newValue': change.new_value,
        'changedAt': change.changed_at.isoformat(),
        'changedBy': change.changed_by
    } for change in changes])

# ========== INQUIRY ENDPOINTS ==========

@app.route('/api/inquiries', methods=['GET'])
def get_inquiries():
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    
    query = Inquiry.query
    
    if search:
        query = query.filter(
            db.or_(
                Inquiry.name.ilike(f'%{search}%'),
                Inquiry.phone.ilike(f'%{search}%'),
                Inquiry.location.ilike(f'%{search}%')
            )
        )
    
    if status:
        query = query.filter(Inquiry.status == status)
    
    inquiries = query.order_by(Inquiry.created_at.desc()).all()
    
    return jsonify([{
        'id': inq.id,
        'name': inq.name,
        'phone': inq.phone,
        'email': inq.email,
        'location': inq.location,
        'status': inq.status,
        'date': inq.date,
        'notes': inq.notes,
        'quoteAmount': inq.quote_amount,
        'createdAt': inq.created_at.isoformat(),
        'updatedAt': inq.updated_at.isoformat()
    } for inq in inquiries])

@app.route('/api/inquiries', methods=['POST'])
def create_inquiry():
    data = request.json
    
    inquiry = Inquiry(
        name=data.get('name'),
        phone=data.get('phone'),
        email=data.get('email'),
        location=data.get('location'),
        status=data.get('status', 'new'),
        date=data.get('date'),
        notes=data.get('notes', ''),
        quote_amount=data.get('quoteAmount')
    )
    
    db.session.add(inquiry)
    db.session.commit()
    
    log_change(Inquiry, inquiry.id, 'created', f'Inquiry from {inquiry.name} created')
    
    return jsonify({'id': inquiry.id, 'message': 'Inquiry created successfully'}), 201

@app.route('/api/inquiries/<int:inquiry_id>', methods=['PUT'])
def update_inquiry(inquiry_id):
    inquiry = Inquiry.query.get_or_404(inquiry_id)
    data = request.json
    
    old_status = inquiry.status
    
    inquiry.name = data.get('name', inquiry.name)
    inquiry.phone = data.get('phone', inquiry.phone)
    inquiry.email = data.get('email', inquiry.email)
    inquiry.location = data.get('location', inquiry.location)
    inquiry.status = data.get('status', inquiry.status)
    inquiry.date = data.get('date', inquiry.date)
    inquiry.notes = data.get('notes', inquiry.notes)
    inquiry.quote_amount = data.get('quoteAmount', inquiry.quote_amount)
    
    db.session.commit()
    
    if old_status != inquiry.status:
        log_change(Inquiry, inquiry.id, 'status_changed',
                  f'Status changed from {old_status} to {inquiry.status}',
                  old_status, inquiry.status)
    else:
        log_change(Inquiry, inquiry.id, 'updated', f'Inquiry for {inquiry.name} updated')
    
    return jsonify({'message': 'Inquiry updated successfully'})

@app.route('/api/inquiries/<int:inquiry_id>', methods=['DELETE'])
def delete_inquiry(inquiry_id):
    inquiry = Inquiry.query.get_or_404(inquiry_id)
    name = inquiry.name
    
    db.session.delete(inquiry)
    db.session.commit()
    
    return jsonify({'message': f'Inquiry from {name} deleted successfully'})

@app.route('/api/inquiries/<int:inquiry_id>/changes', methods=['GET'])
def get_inquiry_changes(inquiry_id):
    changes = InquiryChange.query.filter_by(inquiry_id=inquiry_id).order_by(InquiryChange.changed_at.desc()).all()
    
    return jsonify([{
        'id': change.id,
        'changeType': change.change_type,
        'description': change.change_description,
        'oldValue': change.old_value,
        'newValue': change.new_value,
        'changedAt': change.changed_at.isoformat(),
        'changedBy': change.changed_by
    } for change in changes])

# ========== VEHICLE ENDPOINTS ==========

@app.route('/api/vehicles', methods=['GET'])
def get_vehicles():
    vehicles = Vehicle.query.order_by(Vehicle.created_at.desc()).all()
    
    return jsonify([{
        'id': veh.id,
        'registration': veh.registration,
        'make': veh.make,
        'model': veh.model,
        'colour': veh.colour,
        'fuelType': veh.fuel_type,
        'motDue': veh.mot_due,
        'taxDue': veh.tax_due,
        'insuranceDue': veh.insurance_due,
        'motActioned': veh.mot_actioned,
        'taxActioned': veh.tax_actioned,
        'insuranceActioned': veh.insurance_actioned,
        'createdAt': veh.created_at.isoformat(),
        'updatedAt': veh.updated_at.isoformat()
    } for veh in vehicles])

@app.route('/api/vehicles', methods=['POST'])
def create_vehicle():
    data = request.json
    
    vehicle = Vehicle(
        registration=data.get('registration'),
        make=data.get('make'),
        model=data.get('model'),
        colour=data.get('colour'),
        fuel_type=data.get('fuelType'),
        mot_due=data.get('motDue'),
        tax_due=data.get('taxDue'),
        insurance_due=data.get('insuranceDue'),
        mot_actioned=data.get('motActioned', False),
        tax_actioned=data.get('taxActioned', False),
        insurance_actioned=data.get('insuranceActioned', False)
    )
    
    db.session.add(vehicle)
    db.session.commit()
    
    log_change(Vehicle, vehicle.id, 'created', f'Vehicle {vehicle.registration} added')
    
    return jsonify({'id': vehicle.id, 'message': 'Vehicle created successfully'}), 201

@app.route('/api/vehicles/<int:vehicle_id>', methods=['PUT'])
def update_vehicle(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    data = request.json
    
    vehicle.registration = data.get('registration', vehicle.registration)
    vehicle.make = data.get('make', vehicle.make)
    vehicle.model = data.get('model', vehicle.model)
    vehicle.colour = data.get('colour', vehicle.colour)
    vehicle.fuel_type = data.get('fuelType', vehicle.fuel_type)
    vehicle.mot_due = data.get('motDue', vehicle.mot_due)
    vehicle.tax_due = data.get('taxDue', vehicle.tax_due)
    vehicle.insurance_due = data.get('insuranceDue', vehicle.insurance_due)
    vehicle.mot_actioned = data.get('motActioned', vehicle.mot_actioned)
    vehicle.tax_actioned = data.get('taxActioned', vehicle.tax_actioned)
    vehicle.insurance_actioned = data.get('insuranceActioned', vehicle.insurance_actioned)
    
    db.session.commit()
    
    log_change(Vehicle, vehicle.id, 'updated', f'Vehicle {vehicle.registration} updated')
    
    return jsonify({'message': 'Vehicle updated successfully'})

@app.route('/api/vehicles/<int:vehicle_id>', methods=['DELETE'])
def delete_vehicle(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    registration = vehicle.registration
    
    db.session.delete(vehicle)
    db.session.commit()
    
    return jsonify({'message': f'Vehicle {registration} deleted successfully'})

@app.route('/api/vehicles/<int:vehicle_id>/changes', methods=['GET'])
def get_vehicle_changes(vehicle_id):
    changes = VehicleChange.query.filter_by(vehicle_id=vehicle_id).order_by(VehicleChange.changed_at.desc()).all()
    
    return jsonify([{
        'id': change.id,
        'changeType': change.change_type,
        'description': change.change_description,
        'changedAt': change.changed_at.isoformat()
    } for change in changes])

# ========== INVOICE GENERATION ==========

@app.route('/api/invoice/generate', methods=['POST'])
def generate_invoice_file():
    try:
        data = request.json
        
        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({'error': 'Invoice template not found'}), 500
            
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        ws['B4'] = data.get('invoiceNumber', 'INV-001')
        ws['B5'] = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        ws['B9'] = data.get('clientName', '')
        ws['B10'] = data.get('clientAddress', '')
        ws['B11'] = data.get('clientPhone', '')
        
        import json
        items = json.loads(data.get('items', '[]'))
        start_row = 14
        subtotal = 0
        
        for idx, item in enumerate(items):
            row = start_row + idx
            ws.cell(row=row, column=1).value = idx + 1
            ws.cell(row=row, column=2).value = item.get('description', '')
            ws.cell(row=row, column=3).value = item.get('quantity', 0)
            ws.cell(row=row, column=4).value = item.get('rate', 0)
            amount = item.get('quantity', 0) * item.get('rate', 0)
            ws.cell(row=row, column=5).value = amount
            ws.cell(row=row, column=5).number_format = '£#,##0.00'
            subtotal += amount
        
        vat = subtotal * 0.20
        total = subtotal + vat
        
        ws['E21'] = subtotal
        ws['E21'].number_format = '£#,##0.00'
        ws['E22'] = vat
        ws['E22'].number_format = '£#,##0.00'
        ws['E23'] = total
        ws['E23'].number_format = '£#,##0.00'
        ws['E23'].font = Font(bold=True, size=12)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'invoice_{data.get("invoiceNumber", "001")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ========== MAIN DASHBOARD PAGE ==========

@app.route('/')
def index():
    html_content = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Scaffolding Business Manager</title>
    <script crossorigin src="https://unpkg.com/react@18/umd/react.production.min.js"></script>
    <script crossorigin src="https://unpkg.com/react-dom@18/umd/react-dom.production.min.js"></script>
    <script src="https://unpkg.com/@babel/standalone/babel.min.js"></script>
    <script src="https://cdn.tailwindcss.com"></script>
</head>
<body>
    <div id="root"></div>
    <script type="text/babel">
        const { useState, useEffect } = React;
        const API_URL = 'http://localhost:5000/api';

        function Dashboard() {
            const [activeTab, setActiveTab] = useState('home');
            const [invoices, setInvoices] = useState([]);
            const [inquiries, setInquiries] = useState([]);
            const [vehicles, setVehicles] = useState([]);

            useEffect(() => {
                fetch(API_URL + '/invoices').then(r => r.json()).then(setInvoices).catch(console.error);
                fetch(API_URL + '/inquiries').then(r => r.json()).then(setInquiries).catch(console.error);
                fetch(API_URL + '/vehicles').then(r => r.json()).then(setVehicles).catch(console.error);
            }, []);

            const getTabClass = (tab) => {
                return activeTab === tab 
                    ? 'px-4 py-2 font-semibold text-blue-600 border-b-2 border-blue-600'
                    : 'px-4 py-2 font-semibold text-gray-600';
            };

            const getStatusClass = (status) => {
                if (status === 'paid') return 'px-2 py-1 rounded text-xs bg-green-100 text-green-800';
                return 'px-2 py-1 rounded text-xs bg-yellow-100 text-yellow-800';
            };

            return (
                <div className="min-h-screen bg-gradient-to-br from-blue-50 to-indigo-100">
                    <div className="container mx-auto p-6">
                        <h1 className="text-4xl font-bold text-gray-800 mb-2">🏗️ Scaffolding Business Manager</h1>
                        <p className="text-gray-600 mb-8">Manage invoices, inquiries, and vehicle reminders</p>
                        
                        <div className="grid grid-cols-1 md:grid-cols-3 gap-6 mb-8">
                            <div className="bg-white rounded-lg shadow-lg p-6 border-l-4 border-blue-500">
                                <h3 className="text-gray-600 text-sm font-semibold">Total Invoices</h3>
                                <p className="text-3xl font-bold text-gray-800">{invoices.length}</p>
                            </div>
                            <div className="bg-white rounded-lg shadow-lg p-6 border-l-4 border-green-500">
                                <h3 className="text-gray-600 text-sm font-semibold">New Inquiries</h3>
                                <p className="text-3xl font-bold text-gray-800">{inquiries.filter(i => i.status === 'new').length}</p>
                            </div>
                            <div className="bg-white rounded-lg shadow-lg p-6 border-l-4 border-purple-500">
                                <h3 className="text-gray-600 text-sm font-semibold">Vehicles</h3>
                                <p className="text-3xl font-bold text-gray-800">{vehicles.length}</p>
                            </div>
                        </div>

                        <div className="bg-white rounded-lg shadow-lg p-6">
                            <div className="flex space-x-4 border-b pb-4 mb-6">
                                <button onClick={() => setActiveTab('invoices')} className={getTabClass('invoices')}>
                                    📄 Invoices
                                </button>
                                <button onClick={() => setActiveTab('inquiries')} className={getTabClass('inquiries')}>
                                    👥 Inquiries
                                </button>
                                <button onClick={() => setActiveTab('vehicles')} className={getTabClass('vehicles')}>
                                    🚗 Vehicles
                                </button>
                            </div>

                            {activeTab === 'invoices' && (
                                <div>
                                    <h2 className="text-2xl font-bold mb-4">Invoices</h2>
                                    <p className="text-gray-600 mb-4">Total: {invoices.length} invoices</p>
                                    {invoices.length === 0 ? (
                                        <div className="text-center py-8 text-gray-500">
                                            <p>No invoices yet. Create your first invoice!</p>
                                        </div>
                                    ) : (
                                        <div className="space-y-2">
                                            {invoices.slice(0, 10).map(inv => (
                                                <div key={inv.id} className="p-3 bg-gray-50 rounded border flex justify-between items-center">
                                                    <span className="font-semibold">{inv.invoiceNumber}</span>
                                                    <span>{inv.clientName}</span>
                                                    <span className="text-sm text-gray-600">{inv.date}</span>
                                                    <span className={getStatusClass(inv.status)}>
                                                        {inv.status}
                                                    </span>
                                                </div>
                                            ))}
                                        </div>
                                    )}
                                </div>
                            )}

                            {activeTab === 'inquiries' && (
                                <div>
                                    <h2 className="text-2xl font-bold mb-4">Inquiries</h2>
                                    <p className="text-gray-600 mb-4">Total: {inquiries.length} inquiries</p>
                                    {inquiries.length === 0 ? (
                                        <div className="text-center py-8 text-gray-500">
                                            <p>No inquiries yet. Add your first customer inquiry!</p>
                                        </div>
                                    ) : (
                                        <div className="space-y-2">
                                            {inquiries.slice(0, 10).map(inq => (
                                                <div key={inq.id} className="p-3 bg-gray-50 rounded border flex justify-between items-center">
                                                    <span className="font-semibold">{inq.name}</span>
                                                    <span>{inq.location}</span>
                                                    <span className="text-sm text-gray-600">{inq.phone}</span>
                                                    <span className="text-sm px-2 py-1 bg-blue-100 text-blue-800 rounded">
                                                        {inq.status}
                                                    </span>
                                                </div>
                                            ))}
                                        </div>
                                    )}
                                </div>
                            )}

                            {activeTab === 'vehicles' && (
                                <div>
                                    <h2 className="text-2xl font-bold mb-4">Vehicles</h2>
                                    <p className="text-gray-600 mb-4">Total: {vehicles.length} vehicles</p>
                                    {vehicles.length === 0 ? (
                                        <div className="text-center py-8 text-gray-500">
                                            <p>No vehicles yet. Add your first vehicle!</p>
                                        </div>
                                    ) : (
                                        <div className="space-y-2">
                                            {vehicles.map(veh => (
                                                <div key={veh.id} className="p-3 bg-gray-50 rounded border flex justify-between items-center">
                                                    <span className="font-semibold">{veh.registration}</span>
                                                    <span>{veh.make} {veh.model}</span>
                                                    <span className="text-sm text-gray-600">MOT: {veh.motDue}</span>
                                                    <span className="text-sm text-gray-600">Tax: {veh.taxDue}</span>
                                                </div>
                                            ))}
                                        </div>
                                    )}
                                </div>
                            )}
                        </div>

                        <div className="mt-6 text-center text-gray-500 text-sm">
                            <p>Database: ''' + DB_PATH + '''</p>
                            <p className="mt-2">✅ Server is running. To access full features, open index.html in a browser.</p>
                        </div>
                    </div>
                </div>
            );
        }

        ReactDOM.render(<Dashboard />, document.getElementById('root'));
    </script>
</body>
</html>'''
    return html_content

def init_db():
    with app.app_context():
        db.create_all()
        print("✓ Database initialized successfully!")

def open_browser():
    webbrowser.open('http://localhost:5000')

if __name__ == '__main__':
    print("\n" + "="*60)
    print("  SCAFFOLDING BUSINESS MANAGER")
    print("="*60)
    print(f"\n📊 Database: {DB_PATH}")
    print(f"📄 Template: {TEMPLATE_PATH}")
    print("\nInitializing database...")
    init_db()
    print("\n✓ Server starting on http://localhost:5000")
    print("\n⚠️  KEEP THIS WINDOW OPEN while using the application")
    print("="*60 + "\n")
    
    timer = threading.Timer(1.5, open_browser)
    timer.daemon = True
    timer.start()
    
    app.run(host='0.0.0.0', port=5000, debug=False)